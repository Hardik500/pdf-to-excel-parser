"""
Output generation module for parsed statements.

Supports multiple output formats:
- Excel (XLSX)
- CSV
- JSON
- Normalized transaction list
"""

import csv
import json
from typing import List, Dict, Any, Optional
from datetime import datetime
import os

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


class OutputGenerator:
    """Generate output in various formats from parsed transactions."""

    def __init__(self, transactions: List[Dict[str, Any]], statement_type: str = "bank"):
        """
        Initialize the output generator.

        Args:
            transactions: List of parsed transaction dictionaries
            statement_type: Type of statement ('bank', 'credit_card', 'upi')
        """
        self.transactions = transactions
        self.statement_type = statement_type

    def to_excel(self, filepath: str, include_summary: bool = True) -> bool:
        """
        Export transactions to Excel format.

        Args:
            filepath: Output file path
            include_summary: Whether to include summary sheet

        Returns:
            True if successful, False otherwise
        """
        if not HAS_OPENPYXL:
            raise ImportError(
                "openpyxl is required for Excel output. "
                "Install with: pip install openpyxl"
            )

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Transactions"

        # Determine columns based on statement type
        columns = self._get_columns_for_type()

        # Write header
        header_style = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", fill_type="solid")

        for col_idx, column in enumerate(columns, 1):
            cell = ws.cell(row=1, column=col_idx, value=column)
            cell.font = header_style
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        # Write data
        for row_idx, tx in enumerate(self.transactions, 2):
            for col_idx, column in enumerate(columns, 1):
                value = tx.get(column, '')
                cell = ws.cell(row=row_idx, column=col_idx, value=value)

                # Format numeric columns
                if column in ['amount', 'debit', 'credit', 'balance']:
                    try:
                        cell.value = float(value) if value else 0
                        cell.number_format = '#,##0.00'
                    except (ValueError, TypeError):
                        pass

                # Format date columns
                if column == 'date':
                    try:
                        from datetime import datetime
                        cell.value = datetime.strptime(str(value), "%d/%m/%Y")
                        cell.number_format = "dd/mm/yyyy"
                    except (ValueError, TypeError):
                        pass

        # Auto-adjust column widths
        for col_idx, column in enumerate(columns, 1):
            max_width = len(column) + 2
            for row in ws.iter_rows(min_row=2, max_row=len(self.transactions) + 1, min_col=col_idx, max_col=col_idx):
                for cell in row:
                    if cell.value:
                        max_width = max(max_width, len(str(cell.value)) + 2)
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_width, 50)

        # Add summary sheet if requested
        if include_summary:
            self._add_summary_sheet(wb)

        # Save
        wb.save(filepath)
        return True

    def to_csv(self, filepath: str, delimiter: str = ',') -> bool:
        """
        Export transactions to CSV format.

        Args:
            filepath: Output file path
            delimiter: CSV delimiter (default: comma)

        Returns:
            True if successful
        """
        columns = self._get_columns_for_type()

        with open(filepath, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=columns, delimiter=delimiter,
                                   extrasaction='ignore')
            writer.writeheader()
            writer.writerows(self.transactions)

        return True

    def to_json(self, filepath: str, indent: int = 2) -> bool:
        """
        Export transactions to JSON format.

        Args:
            filepath: Output file path
            indent: JSON indentation level

        Returns:
            True if successful
        """
        data = {
            'generated_at': datetime.now().isoformat(),
            'statement_type': self.statement_type,
            'transaction_count': len(self.transactions),
            'transactions': self.transactions,
        }

        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(data, f, indent=indent, default=str)

        return True

    def to_list(self) -> List[Dict[str, Any]]:
        """
        Return transactions as a list of dictionaries.

        Returns:
            List of transaction dictionaries
        """
        return self.transactions

    def _get_columns_for_type(self) -> List[str]:
        """Get appropriate columns for the statement type."""
        column_maps = {
            'bank': [
                'date', 'narration', 'value_date', 'debit', 'credit',
                'balance', 'reference'
            ],
            'credit_card': [
                'date', 'merchant', 'amount', 'type', 'card_no', 'reference'
            ],
            'upi': [
                'date', 'merchant', 'amount', 'type', 'reference', 'upi_ref'
            ],
            'default': [
                'date', 'description', 'amount', 'type', 'reference'
            ],
        }

        return column_maps.get(self.statement_type, column_maps['default'])

    def _add_summary_sheet(self, wb):
        """Add a summary sheet to the workbook."""
        summary = wb.create_sheet("Summary")

        # Calculate summary statistics
        total_credit = sum(float(tx.get('credit', 0) or 0) for tx in self.transactions)
        total_debit = sum(float(tx.get('debit', 0) or 0) for tx in self.transactions)
        net_amount = total_credit - total_debit

        # Header
        summary.cell(row=1, column=1, value="Summary")
        summary.cell(row=1, column=1).font = Font(bold=True, size=14)

        # Stats
        summary.cell(row=3, column=1, value="Total Transactions")
        summary.cell(row=3, column=2, value=len(self.transactions))

        summary.cell(row=4, column=1, value="Total Credits (Income)")
        summary.cell(row=4, column=2, value=total_credit)
        summary.cell(row=4, column=2).number_format = '#,##0.00'

        summary.cell(row=5, column=1, value="Total Debits (Expenses)")
        summary.cell(row=5, column=2, value=total_debit)
        summary.cell(row=5, column=2).number_format = '#,##0.00'

        summary.cell(row=6, column=1, value="Net Amount")
        summary.cell(row=6, column=2, value=net_amount)
        summary.cell(row=6, column=2).number_format = '#,##0.00'

        # Color code
        net_cell = summary.cell(row=6, column=2)
        if net_amount > 0:
            net_cell.fill = PatternFill(start_color="C6EFCE", fill_type="solid")
        elif net_amount < 0:
            net_cell.fill = PatternFill(start_color="FFC7CE", fill_type="solid")

        # Auto-adjust
        for col in [1, 2]:
            max_width = 20
            summary.column_dimensions[get_column_letter(col)].width = max_width


def generate_output(transactions: List[Dict[str, Any]], statement_type: str = "bank",
                   format: str = "excel", filepath: str = None) -> Optional[str]:
    """
    Generate output in the specified format.

    Args:
        transactions: List of transaction dictionaries
        statement_type: Type of statement
        format: Output format ('excel', 'csv', 'json')
        filepath: Output file path (required for file outputs)

    Returns:
        JSON string for json format, None otherwise
    """
    generator = OutputGenerator(transactions, statement_type)

    if format == 'excel':
        if not filepath:
            raise ValueError("filepath is required for Excel output")
        generator.to_excel(filepath)
        return None
    elif format == 'csv':
        if not filepath:
            raise ValueError("filepath is required for CSV output")
        generator.to_csv(filepath)
        return None
    elif format == 'json':
        if not filepath:
            return json.dumps({
                'transactions': transactions,
                'statement_type': statement_type,
                'count': len(transactions),
            }, default=str)
        generator.to_json(filepath)
        return None
    else:
        raise ValueError(f"Unknown format: {format}. Use 'excel', 'csv', or 'json'.")
